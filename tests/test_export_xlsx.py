import atexit
import json
import os
import tempfile
import unittest
from unittest import mock

from openpyxl import load_workbook


TEST_DATA_DIR = tempfile.TemporaryDirectory()
atexit.register(TEST_DATA_DIR.cleanup)
os.environ["CRM_DATA_DIR"] = TEST_DATA_DIR.name
os.environ["CRM_DESKTOP_APP"] = "0"

import app as app_module


class ExportXlsxTest(unittest.TestCase):
    def setUp(self):
        self.output_dir = tempfile.TemporaryDirectory()
        self.original_barcode_dir = app_module.BARCODE_DIR
        app_module.BARCODE_DIR = self.output_dir.name
        app_module.app.config.update(TESTING=True)
        self.client = app_module.app.test_client()
        response = self.client.post(
            "/api/app-auth/login",
            json={"username": "admin", "password": "88293529"},
        )
        self.assertTrue(response.get_json()["success"])

    def tearDown(self):
        app_module.BARCODE_DIR = self.original_barcode_dir
        self.output_dir.cleanup()

    def test_export_removes_characters_forbidden_in_xlsx_cells(self):
        response = self.client.post("/api/export/xlsx", json={
            "barcodes": [{
                "barcode": "TEST001",
                "time": "2026-07-20 12:00:00",
                "remark": "客户\x00备\x0b注\x1f",
                "fields": {"sr2": {"name1": "客\x0c户"}},
            }]
        })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])

        workbook = load_workbook(os.path.join(self.output_dir.name, "export_result.xlsx"))
        self.assertEqual(workbook.active.cell(row=3, column=2).value, "客户备注")
        self.assertEqual(workbook.active.cell(row=3, column=4).value, "客户")
        workbook.close()

    def test_exported_xlsx_is_served_as_a_download(self):
        export_response = self.client.post("/api/export/xlsx", json={
            "barcodes": [{"barcode": "TEST002", "fields": {}}]
        })
        self.assertTrue(export_response.get_json()["success"])

        with self.client.get("/barcode/export_result.xlsx") as download_response:
            self.assertEqual(download_response.status_code, 200)
            self.assertEqual(
                download_response.mimetype,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.assertIn("attachment", download_response.headers["Content-Disposition"])

    def test_service_order_export_uses_install_template_columns_and_deduplicates(self):
        service_dir = tempfile.TemporaryDirectory()
        self.addCleanup(service_dir.cleanup)
        detail = {
            "service_no": "FWD202608010001",
            "fields": [
                {"label": "客户姓名", "value": "王剑利"},
                {"label": "联系电话", "value": "18679402783"},
                {"label": "联系地址", "value": "江西省 抚州市 南城县 洪卫壹品9-1601"},
                {"label": "受理时间", "value": "2026-08-01 09:00"},
                {"label": "客户预约时间", "value": "2026-08-03 10:00"},
                {"label": "服务人员", "value": "张工"},
            ],
            "products": [{"product_name": "壁挂式饮水机（白色）EWD600S", "product_code": "906020905", "barcode": "5312408100080"}],
        }
        with open(os.path.join(service_dir.name, "FWD202608010001.json"), "w", encoding="utf-8") as file:
            json.dump(detail, file, ensure_ascii=False)
        items = [
            {"barcode": "TEST001", "fields": {"sr2": [{"servno1": "FWD202608010001", "typestr1": "安装", "servdate1": "2026-08-01"}]}},
            {"barcode": "TEST002", "fields": {"sr2": [{"servno1": "FWD202608010001", "typestr1": "安装", "servdate1": "2026-08-01"}]}},
            {"barcode": "TEST003", "fields": {"sr2": [{"servno1": "FWD202608020001", "typestr1": "安装", "servdate1": "2026-08-02"}]}},
        ]
        with mock.patch.object(app_module, "SERVICE_ORDER_DIR", service_dir.name), mock.patch.object(app_module, "scan_barcodes", return_value=items):
            response = self.client.post("/api/service-orders/export/xlsx", json={"barcodes": ["TEST001", "TEST002", "TEST003"]})

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["exported_count"], 1)
        self.assertEqual(payload["skipped_count"], 1)
        workbook = load_workbook(os.path.join(self.output_dir.name, payload["filename"]))
        sheet = workbook.active
        self.assertEqual(sheet.title, "产品销售安装")
        self.assertEqual(sheet["B1"].value, "用户姓名")
        self.assertEqual(sheet["U1"].value, "产品类型")
        self.assertEqual(sheet["B2"].value, "产品销售安装")
        self.assertEqual(sheet["B3"].value, "王剑利")
        self.assertEqual(sheet["E3"].value, "2026-08-01 09:00")
        self.assertEqual(sheet["F3"].value, "服务单号：FWD202608010001")
        self.assertEqual(sheet["I3"].value, "906020905")
        self.assertEqual(sheet["J3"].value, "EWD600S")
        self.assertEqual(sheet["L3"].value, "张工")
        self.assertEqual(sheet["M3"].value, "5312408100080")
        self.assertEqual(sheet["N3"].value, "2026-08-03 10:00")
        workbook.close()


if __name__ == "__main__":
    unittest.main()
