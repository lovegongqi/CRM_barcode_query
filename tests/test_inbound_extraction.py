import unittest

from openpyxl import load_workbook

from inbound_extraction import (
    build_inbound_result,
    build_inbound_workbook,
    normalize_packing_slip_no,
)


PACKING_SLIP_NO = "SH202607210002"


class InboundExtractionTest(unittest.TestCase):
    def setUp(self):
        self.rows = [
            {
                "page": 1,
                "row_index": 1,
                "order_number": "210524",
                "product_code": "916000024",
                "description": "中央净水机",
                "expected_quantity": "2",
                "serial": "SN00000001",
            },
            {
                "page": 1,
                "row_index": 2,
                "order_number": "210524",
                "product_code": "916000024",
                "description": "中央净水机",
                "expected_quantity": "2",
                "serial": "SN00000002",
            },
            {
                "page": 2,
                "row_index": 1,
                "order_number": "210525",
                "product_code": "916000025",
                "description": "软水机",
                "expected_quantity": "2",
                "serial": "SN00000002",
            },
            {
                "page": 2,
                "row_index": 2,
                "order_number": "210525",
                "product_code": "916000025",
                "description": "软水机",
                "expected_quantity": "2",
                "serial": "SN00000003",
            },
        ]
        self.page_counts = [
            {"page": 1, "row_count": 2},
            {"page": 2, "row_count": 2},
        ]

    def test_normalizes_packing_slip_number(self):
        self.assertEqual(
            normalize_packing_slip_no(" sh202607210002 "),
            PACKING_SLIP_NO,
        )

    def test_rejects_invalid_packing_slip_number(self):
        with self.assertRaisesRegex(ValueError, "装箱单号格式不正确"):
            normalize_packing_slip_no("210524")

    def test_builds_grouped_result_with_global_duplicate_reporting(self):
        result = build_inbound_result(
            PACKING_SLIP_NO,
            self.rows,
            self.page_counts,
            packing_slip_type="销售订单",
        )

        self.assertEqual(result["packing_slip_type"], "销售订单")
        self.assertEqual(result["pages_read"], [1, 2])
        self.assertEqual(result["total_serials"], 3)
        self.assertEqual(result["duplicate_serials"], ["SN00000002"])
        self.assertEqual(
            [item["product_code"] for item in result["items"]],
            ["916000024", "916000025"],
        )
        self.assertEqual(result["items"][0]["order_numbers"], ["210524"])
        self.assertTrue(result["items"][1]["quantity_mismatch"])

    def test_workbook_has_one_row_per_clean_serial_in_source_order(self):
        result = build_inbound_result(
            PACKING_SLIP_NO,
            self.rows,
            self.page_counts,
        )

        workbook = load_workbook(build_inbound_workbook(result))
        sheet = workbook.active
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["装箱单号", "订单号", "物料编码", "物料描述", "应发数量", "条码", "明细类型", "来源页码"],
        )
        self.assertEqual(sheet.max_row, result["total_serials"] + 1)
        self.assertEqual([sheet["F2"].value, sheet["F3"].value], ["SN00000001", "SN00000002"])
        self.assertIsInstance(sheet["H2"].value, int)
        workbook.close()

    def test_keeps_unbarcoded_accessory_from_shipment_details(self):
        result = build_inbound_result(
            PACKING_SLIP_NO,
            [
                {
                    "page": 1,
                    "row_index": 1,
                    "product_code": "906042856",
                    "description": "净水机",
                    "serial": "9462607180448",
                }
            ],
            [{"page": 1, "row_count": 1}],
            shipment_rows=[
                {"product_code": "906042856", "description": "净水机", "expected_quantity": "1"},
                {"product_code": "247296319", "description": "中央净水机面贴", "expected_quantity": "10"},
            ],
        )

        self.assertEqual([item["product_code"] for item in result["items"]], ["906042856", "247296319"])
        accessory = result["items"][1]
        self.assertEqual(accessory["expected_quantity"], 10)
        self.assertEqual(accessory["serial_count"], 0)
        self.assertEqual(accessory["unbarcoded_quantity"], 10)
        self.assertEqual(result["total_serials"], 1)
        self.assertEqual(result["rows"][-1]["serial"], "")
        self.assertEqual(result["rows"][-1]["record_type"], "无条码配件")


if __name__ == "__main__":
    unittest.main()
